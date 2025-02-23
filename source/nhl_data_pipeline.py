"""
Module to scrape data from a website and save it to an Excel file.
"""
import asyncio
import os
import time
import urllib.parse
import zipfile

import aiohttp
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# Constants
BASE_URL = "https://www.scrapethissite.com/pages/forms/"
OUTPUT_ZIP = "output/scraped_pages.zip"
EXCEL_FILE = "output/NHL_Stats.xlsx"
SHEET_1_NAME = "NHL Stats 1990-2011"
SHEET_2_NAME = "Winner and Loser per Year"


async def fetch_page(session, url):
    """
    Fetch a page and return its HTML content.
    :param session:
    :type session:
    :param url:
    :type url:
    :return:
    :rtype:
    """
    async with session.get(url) as response:
        return await response.text()


async def get_pages_url(session):
    """
    Find the total number of pages by extracting pagination links.
    :param session:
    :type session:
    :return:
    :rtype:
    """
    page_urls = []
    try:
        html = await fetch_page(session, BASE_URL)
        soup = BeautifulSoup(html, "html.parser")

        # Find the pagination section
        pagination = soup.find("ul", class_="pagination")

        # Get page urls by excluding the 'Next' and 'Previous' links using 'aria-label' attribute
        if pagination:
            page_urls = [a['href'] for a in pagination.find_all("a", attrs={"aria-label": False})]
    except Exception as exc:
        print(f"An error occurred: {exc}")

    return page_urls


async def extract():
    """
    Scrape all pages and save their HTML content to a ZIP archive.
    :return:
    :rtype:
    """
    try:
        async with aiohttp.ClientSession() as session:
            page_urls = await get_pages_url(session)
            print(f"Total Pages Found: {page_urls}")

            # Fetch all pages asynchronously
            tasks = [fetch_page(session, urllib.parse.urljoin(BASE_URL, page_url)) for page_url in page_urls]
            results = await asyncio.gather(*tasks)

            rows = []
            headers = []
            # Save all HTML pages inside a ZIP archive
            with zipfile.ZipFile(OUTPUT_ZIP, "w") as zip_file:
                for page_num, html in enumerate(results, start=1):
                    rows, headers = get_html_table(html, rows)
                    save_html_to_zip(zip_file, html, page_num)

            # Save scraped data into an Excel file
            save_to_excel(rows, headers)
            print(f"Scraped {page_num} pages and saved to {OUTPUT_ZIP}")
    except Exception as exc:
        print(f"An error occurred: {exc}")


def get_html_table(html, rows):
    """
    Fetch table data
    :param html:
    :param rows:
    """
    try:
        soup = BeautifulSoup(html, "html.parser")

        # Find the table
        table = soup.find("table", class_="table")

        # Extract headers
        headers = [th.text.strip() for th in table.find_all("th")]

        # Extract rows after Skipping header row
        for tr in table.find_all("tr")[1:]:
            cells = [td.text.strip() for td in tr.find_all("td")]
            # Skip empty rows
            if cells:
                rows.append(cells)
        return rows, headers
    except Exception as exc:
        print(f"An error occurred: {exc}")


def save_html_to_zip(zip_file, html, page_num):
    """
    Save HTML pages inside a ZIP archive.
    :param zip_file:
    :param html:
    :param page_num:
    """
    try:
        file_name = f"{page_num}.html"
        with open(file_name, "w", encoding="utf-8") as f:
            f.write(html)
        zip_file.write(file_name)
        os.remove(file_name)
    except Exception as exc:
        print(f"An error occurred: {exc}")


def save_to_excel(rows, headers):
    """
    Save scraped data into an Excel file.
    :param rows:
    :param headers:
    """
    try:
        # Create an Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_1_NAME

        # Write headers
        ws.append(headers)

        # Write data rows
        for row in rows:
            ws.append(row)

        # Save Excel file
        wb.save(EXCEL_FILE)
        print(f"Excel file '{EXCEL_FILE}' created successfully with sheet '{SHEET_1_NAME}'")
    except Exception as exc:
        print(f"An error occurred: {exc}")


def transform():
    """
    Transform scraped data into an Excel file.
    """
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_1_NAME]

        # Read data into a dictionary {year: {team: num_wins}}
        stats = {}

        # Identify column indexes
        header = [cell.value for cell in ws[1]]
        year_idx = header.index("Year")
        team_idx = header.index("Team Name")
        wins_idx = header.index("Wins")

        # Process rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            year, team, wins = row[year_idx], row[team_idx], row[wins_idx]

            if year not in stats:
                stats[year] = {}

            # {'1990': {'Boston Bruins': '44', 'Buffalo Sabres': '31'}}
            stats[year][team] = wins

        # Compute winners and losers
        sheet_2_headers = [["Year", "Winner", "Winner Num. of Wins", "Loser", "Loser Num. of Wins"]]

        for year, teams in stats.items():
            # Team with most wins
            winner = max(teams, key=teams.get)
            # Team with the least wins
            loser = min(teams, key=teams.get)
            sheet_2_headers.append([year, winner, teams[winner], loser, teams[loser]])

        # Create a new sheet and write data
        if SHEET_2_NAME in wb.sheetnames:
            wb.remove(wb[SHEET_2_NAME])

        ws_summary = wb.create_sheet(SHEET_2_NAME)

        for row in sheet_2_headers:
            ws_summary.append(row)

        # Save the workbook
        wb.save(EXCEL_FILE)
        print(f"Sheet '{SHEET_2_NAME}' created successfully.")
    except Exception as exc:
        print(f"An error occurred: {exc}")


if __name__ == "__main__":
    start_time = time.time()
    asyncio.run(extract())
    transform()
    print(f"Execution time: {time.time() - start_time} seconds.")
